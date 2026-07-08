import io
import re
import datetime
import unicodedata
import openpyxl

def normalize_text(text):
    """
    Normalizes header titles: converts to lowercase, removes accents,
    strips whitespace, and consolidates internal spaces.
    """
    if text is None:
        return ""
    # Convert to string and strip
    val = str(text).strip().lower()
    # Normalize unicode to remove accents/diacritics
    val = unicodedata.normalize('NFKD', val).encode('ASCII', 'ignore').decode('ASCII')
    # Remove excessive whitespace
    val = re.sub(r'[\r\n\t\s]+', ' ', val)
    return val.strip()

def find_headers(sheet):
    """
    Scans the first few rows of a worksheet to find where all 5 mandatory columns are present.
    Returns a dictionary mapping 'nome', 'funcao', 'data', 'cid', 'quantidade', and optionally 'inss'
    to cell indices (0-indexed), as well as the 1-based index of the header row. Returns None if not found.
    """
    for r_idx, row in enumerate(sheet.iter_rows(max_row=10, values_only=True), 1):
        if not any(row):
            continue
        mapping = {}
        for c_idx, val in enumerate(row):
            if val is None:
                continue
            norm = normalize_text(val)
            # Matching Nome
            if "nome" in norm and "colaborador" in norm:
                mapping["nome"] = c_idx
            elif norm == "nome":
                mapping["nome"] = c_idx
            # Matching Função
            elif "funcao" in norm or "cargo" in norm:
                mapping["funcao"] = c_idx
            # Matching Data
            elif "data" in norm and "inicial" in norm:
                mapping["data"] = c_idx
            elif norm == "data":
                mapping["data"] = c_idx
            # Matching CID
            elif norm == "cid":
                mapping["cid"] = c_idx
            # Matching Quantidade
            elif "quantidade" in norm or "dias" in norm or "horas" in norm:
                mapping["quantidade"] = c_idx
            # Matching INSS (optional)
            elif "inss" in norm:
                mapping["inss"] = c_idx

        # We need the 5 mandatory mappings to consider it a valid header row
        mandatory = {"nome", "funcao", "data", "cid", "quantidade"}
        if mandatory.issubset(mapping.keys()):
            return mapping, r_idx
    return None, None

def parse_quantity(val, default_unit_behavior="reject"):
    """
    Parses the Quantity column. Handles numeric values and formats like:
    '2 dias', '1 dia', '4 horas', '4h', '8:00'.
    Returns (numeric_value, unit) where unit is 'days' or 'hours'.
    Returns (None, None) if parsing fails or unit behavior dictates rejection.
    """
    if val is None:
        return None, None

    # Handle float/int directly
    if isinstance(val, (int, float)):
        val_num = float(val)
        if default_unit_behavior == "reject":
            return None, None
        elif default_unit_behavior == "days":
            return val_num, "days"
        elif default_unit_behavior == "hours":
            return val_num, "hours"
        return None, None

    val_str = str(val).strip().lower()
    if not val_str:
        return None, None

    # Handle Time Formats like 8:00, 08:30
    time_match = re.match(r'^(\d{1,2}):(\d{2})(?::\d{2})?$', val_str)
    if time_match:
        h = int(time_match.group(1))
        m = int(time_match.group(2))
        hours_val = h + (m / 60.0)
        return hours_val, "hours"

    # Match digits followed by optional unit
    match = re.match(r'^([\d.,]+)\s*(dia[s]?|d|hora[s]?|h)?$', val_str)
    if match:
        num_str = match.group(1).replace(",", ".")
        try:
            num = float(num_str)
        except ValueError:
            return None, None

        unit_str = match.group(2)
        if unit_str in ("dia", "dias", "d"):
            return num, "days"
        if unit_str in ("hora", "horas", "h"):
            return num, "hours"
        if not unit_str:
            # Numeric value without unit
            if default_unit_behavior == "reject":
                return None, None
            if default_unit_behavior == "days":
                return num, "days"
            if default_unit_behavior == "hours":
                return num, "hours"

    return None, None

def parse_date(val):
    """
    Parses dates from Excel. Can be datetime object or string.
    Returns datetime.date object, or None if invalid.
    """
    if isinstance(val, (datetime.date, datetime.datetime)):
        if isinstance(val, datetime.datetime):
            return val.date()
        return val

    if val is None:
        return None

    val_str = str(val).strip()
    if not val_str:
        return None

    # Try common formats
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.datetime.strptime(val_str, fmt).date()
        except ValueError:
            continue

    return None

def normalize_inss_value(val):
    """
    Normalizes the INSS column values.
    Returns (normalized_value, warning_msg).
    Normalized values: 'SIM', 'NÃO', 'NÃO INFORMADO'.
    """
    if val is None:
        return "NÃO INFORMADO", None
    
    val_str = str(val).strip().lower()
    # Normalize unicode to remove accents (e.g. não -> nao)
    val_norm = unicodedata.normalize('NFKD', val_str).encode('ASCII', 'ignore').decode('ASCII')
    
    if val_norm in ("sim", "s"):
        return "SIM", None
    if val_norm in ("nao", "n"):
        return "NÃO", None
    if not val_str:
        return "NÃO INFORMADO", None
    
    # Invalid value, returns warning message and treats as NÃO INFORMADO
    return "NÃO INFORMADO", f"Valor '{val}' inválido na coluna do INSS. Tratado como NÃO INFORMADO."

def parse_excel_in_memory(file_content, default_unit_behavior="reject", max_rows=1000):
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True, read_only=True)
    except Exception:
        return {
            "success": False,
            "error_msg": "Arquivo corrompido ou formato inválido. Apenas planilhas .xlsx válidas são aceitas."
        }

    # Find the first compatible sheet
    active_sheet = None
    header_mapping = None
    header_row_idx = None

    for sheet in wb.worksheets:
        mapping, r_idx = find_headers(sheet)
        if mapping:
            active_sheet = sheet
            header_mapping = mapping
            header_row_idx = r_idx
            break

    if not active_sheet:
        return {
            "success": False,
            "error_msg": "Planilha incompatível. Certifique-se de que o arquivo contém as colunas obrigatórias: 'Nome do colaborador', 'Função do colaborador', 'Data inicial do atestado', 'CID' e 'Quantidade de dias/ HORAS'."
        }

    rows_parsed = 0
    valid_records = []
    inconsistencies = []
    duplicates_list = []
    seen_tuples = set()
    duplicate_count = 0

    idx_nome = header_mapping["nome"]
    idx_funcao = header_mapping["funcao"]
    idx_data = header_mapping["data"]
    idx_cid = header_mapping["cid"]
    idx_qty = header_mapping["quantidade"]
    idx_inss = header_mapping.get("inss")
    inss_available = idx_inss is not None

    # Iterate rows starting after the header row
    for curr_row_idx, row in enumerate(active_sheet.iter_rows(min_row=header_row_idx + 1, values_only=True), header_row_idx + 1):
        # Skip completely empty rows
        if not any(v is not None for v in row):
            continue

        rows_parsed += 1
        if rows_parsed > max_rows:
            inconsistencies.append({
                "linha": curr_row_idx,
                "erro": f"Limite de linhas excedido (máximo {max_rows} linhas). Esta linha e as seguintes foram desconsideradas.",
                "detalhe": ""
            })
            break

        # Extract cells safely
        raw_nome = row[idx_nome] if idx_nome < len(row) else None
        raw_funcao = row[idx_funcao] if idx_funcao < len(row) else None
        raw_data = row[idx_data] if idx_data < len(row) else None
        raw_cid = row[idx_cid] if idx_cid < len(row) else None
        raw_qty = row[idx_qty] if idx_qty < len(row) else None

        # 1. Validate Nome
        if raw_nome is None or not str(raw_nome).strip():
            inconsistencies.append({
                "linha": curr_row_idx,
                "erro": "Nome do colaborador ausente",
                "detalhe": "A coluna do nome não pode estar em branco."
            })
            continue
        nome = str(raw_nome).strip()
        # Clean double/excess spaces
        nome = re.sub(r'\s+', ' ', nome)

        # 2. Validate Função
        funcao = str(raw_funcao).strip() if raw_funcao is not None else ""
        funcao = re.sub(r'\s+', ' ', funcao)
        if not funcao:
            funcao = "Não informada"

        # 3. Validate Data
        data_parsed = parse_date(raw_data)
        if data_parsed is None:
            inconsistencies.append({
                "linha": curr_row_idx,
                "erro": "Data de atestado inválida ou ausente",
                "detalhe": f"Valor '{raw_data}' não pôde ser convertido em data."
            })
            continue

        # 4. Validate Quantity and Unit
        qty_num, unit = parse_quantity(raw_qty, default_unit_behavior)
        if qty_num is None:
            inconsistencies.append({
                "linha": curr_row_idx,
                "erro": "Quantidade de dias/horas inválida ou sem unidade",
                "detalhe": f"Valor '{raw_qty}' não pôde ser analisado conforme regra selecionada."
            })
            continue

        # 5. CID
        cid = str(raw_cid).strip().upper() if raw_cid is not None else ""
        cid = re.sub(r'\s+', '', cid)
        if not cid:
            cid = "Não informado"

        # 6. INSS (optional)
        inss_val = "NÃO INFORMADO"
        if inss_available:
            raw_inss = row[idx_inss] if idx_inss < len(row) else None
            inss_val, warning = normalize_inss_value(raw_inss)
            if warning:
                inconsistencies.append({
                    "linha": curr_row_idx,
                    "erro": "Valor inválido na coluna do INSS",
                    "detalhe": warning
                })

        # Check for potential duplicates
        record_tuple = (nome, funcao, data_parsed.isoformat(), cid, qty_num, unit, inss_val)
        if record_tuple in seen_tuples:
            duplicate_count += 1
            duplicates_list.append({
                "linha": curr_row_idx,
                "colaborador": nome,
                "data": data_parsed.strftime("%d/%m/%Y"),
                "cid": cid,
                "afastamento": f"{qty_num} { 'dia(s)' if unit == 'days' else 'hora(s)' }"
            })
        else:
            seen_tuples.add(record_tuple)

        # Add to valid records
        valid_records.append({
            "colaborador": nome,
            "funcao": funcao,
            "data": data_parsed,
            "cid": cid,
            "quantidade": qty_num,
            "unidade": unit,
            "inss": inss_val
        })

    # Perform Aggregations
    total_valid = len(valid_records)
    total_invalid = len(inconsistencies)
    colaboradores_unicos = len(set(r["colaborador"] for r in valid_records))
    
    total_dias = sum(r["quantidade"] for r in valid_records if r["unidade"] == "days")
    total_horas = sum(r["quantidade"] for r in valid_records if r["unidade"] == "hours")

    min_date = None
    max_date = None
    if valid_records:
        dates = [r["data"] for r in valid_records]
        min_date = min(dates)
        max_date = max(dates)

    periodo_str = "N/A"
    if min_date and max_date:
        periodo_str = f"{min_date.strftime('%d/%m/%Y')} até {max_date.strftime('%d/%m/%Y')}"

    # INSS Aggregations
    inss_sim_count = 0
    inss_nao_count = 0
    inss_nao_informado_count = 0
    if inss_available:
        for r in valid_records:
            if r["inss"] == "SIM":
                inss_sim_count += 1
            elif r["inss"] == "NÃO":
                inss_nao_count += 1
            else:
                inss_nao_informado_count += 1

    percentage_sim = 0.0
    denom = inss_sim_count + inss_nao_count
    if denom > 0:
        percentage_sim = (inss_sim_count / denom) * 100.0

    # Charts Data Preparation
    # 1. Atestados por Colaborador
    atestados_por_colab = {}
    # Colaborador summaries for summary_table
    colab_summaries = {}

    for r in valid_records:
        colab = r["colaborador"]
        qty = r["quantidade"]
        unit = r["unidade"]
        date = r["data"]
        inss = r["inss"]

        atestados_por_colab[colab] = atestados_por_colab.get(colab, 0) + 1
        
        if colab not in colab_summaries:
            colab_summaries[colab] = {
                "colaborador": colab,
                "atestados_count": 0,
                "total_dias": 0.0,
                "total_horas": 0.0,
                "first_date": date,
                "last_date": date,
                "inss_sim_count": 0,
                "inss_nao_count": 0,
                "inss_nao_informado_count": 0
            }
        
        cs = colab_summaries[colab]
        cs["atestados_count"] += 1
        if date < cs["first_date"]:
            cs["first_date"] = date
        if date > cs["last_date"]:
            cs["last_date"] = date

        if inss == "SIM":
            cs["inss_sim_count"] += 1
        elif inss == "NÃO":
            cs["inss_nao_count"] += 1
        else:
            cs["inss_nao_informado_count"] += 1

        if unit == "days":
            cs["total_dias"] += qty
        elif unit == "hours":
            cs["total_horas"] += qty

    # Sort Atestados per Colab descending
    chart1_labels = sorted(atestados_por_colab.keys(), key=lambda k: atestados_por_colab[k], reverse=True)
    chart1_values = [atestados_por_colab[k] for k in chart1_labels]

    # 2. Evolução temporal (atestados por mês)
    atestados_por_mes = {}
    for r in valid_records:
        date = r["data"]
        key = (date.year, date.month)
        atestados_por_mes[key] = atestados_por_mes.get(key, 0) + 1

    sorted_temporal_keys = sorted(atestados_por_mes.keys())
    portuguese_months = {
        1: "Jan", 2: "Fev", 3: "Mar", 4: "Abr", 5: "Mai", 6: "Jun",
        7: "Jul", 8: "Ago", 9: "Set", 10: "Out", 11: "Nov", 12: "Dez"
    }
    chart2_labels = [f"{portuguese_months[k[1]]}/{k[0]}" for k in sorted_temporal_keys]
    chart2_values = [atestados_por_mes[k] for k in sorted_temporal_keys]

    # 3. Atestados por Função
    atestados_por_funcao = {}
    for r in valid_records:
        func = r["funcao"]
        atestados_por_funcao[func] = atestados_por_funcao.get(func, 0) + 1

    chart3_labels = sorted(atestados_por_funcao.keys(), key=lambda k: atestados_por_funcao[k], reverse=True)
    chart3_values = [atestados_por_funcao[k] for k in chart3_labels]

    # 4. Distribuição agregada por CID (Top 7 + Outros)
    atestados_por_cid = {}
    for r in valid_records:
        c = r["cid"]
        atestados_por_cid[c] = atestados_por_cid.get(c, 0) + 1

    sorted_cids = sorted(atestados_por_cid.keys(), key=lambda k: atestados_por_cid[k], reverse=True)
    chart6_labels = []
    chart6_values = []
    
    if len(sorted_cids) <= 8:
        for c in sorted_cids:
            chart6_labels.append(c)
            chart6_values.append(atestados_por_cid[c])
    else:
        top_cids = sorted_cids[:7]
        other_sum = sum(atestados_por_cid[c] for c in sorted_cids[7:])
        for c in top_cids:
            chart6_labels.append(c)
            chart6_values.append(atestados_por_cid[c])
        chart6_labels.append("Outros")
        chart6_values.append(other_sum)

    # Format summary table data
    summary_table = []
    for k in sorted(colab_summaries.keys()):
        cs = colab_summaries[k]
        summary_table.append({
            "colaborador": cs["colaborador"],
            "atestados_count": cs["atestados_count"],
            "total_dias": cs["total_dias"],
            "total_horas": cs["total_horas"],
            "first_date": cs["first_date"].strftime("%d/%m/%Y"),
            "last_date": cs["last_date"].strftime("%d/%m/%Y"),
            "inss_sim_count": cs["inss_sim_count"],
            "inss_nao_count": cs["inss_nao_count"],
            "inss_nao_informado_count": cs["inss_nao_informado_count"]
        })

    # Prepare data dict to return
    result = {
        "success": True,
        "sheet_name": active_sheet.title,
        "total_rows": rows_parsed,
        "valid_count": total_valid,
        "invalid_count": total_invalid,
        "duplicate_count": duplicate_count,
        "inss_available": inss_available,
        "indicators": {
            "total_valid": total_valid,
            "colab_count": colaboradores_unicos,
            "total_dias": total_dias,
            "total_horas": total_horas,
            "total_invalid": total_invalid,
            "total_duplicates": duplicate_count,
            "periodo": periodo_str,
            "inss_sim_count": inss_sim_count,
            "inss_nao_count": inss_nao_count,
            "inss_nao_informado_count": inss_nao_informado_count,
            "inss_percentage_sim": percentage_sim
        },
        "charts_data": {
            "colab_atestados": {"labels": chart1_labels, "values": chart1_values},
            "temporal": {"labels": chart2_labels, "values": chart2_values},
            "funcao": {"labels": chart3_labels, "values": chart3_values},
            "cid": {"labels": chart6_labels, "values": chart6_values}
        },
        "summary_table": summary_table,
        "inconsistencies": inconsistencies,
        "duplicates_list": duplicates_list,
        "raw_valid_records": [
            {
                "colaborador": r["colaborador"],
                "funcao": r["funcao"],
                "data": r["data"].strftime("%Y-%m-%d"),
                "cid": r["cid"],
                "quantidade": r["quantidade"],
                "unidade": r["unidade"],
                "inss": r["inss"]
            }
            for r in valid_records
        ]
    }
    return result
